from sendgrid.helpers.mail import Mail, Email, To, Content
from sendgrid import SendGridAPIClient
from datetime import datetime, timedelta
from googleapiclient.discovery import build
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter.filedialog import askopenfile
from jinja2 import Template
import urllib.request
import tkinter
import customtkinter
import pandas as pd
import re

#Function to calculate the date to look for
def CalculateDate():
    actual_time = datetime.utcnow()
    date_tolook = actual_time - \
                        timedelta(days = maxdate)
    date_tolook = date_tolook.isoformat("T") +"Z"
    return date_tolook

# Functio to expand the width of cells in relation with the number of characters
def expandCsvCells(ws):
    column_widths = []
    for row in ws:
        for i,cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                     column_widths += [len(str(cell.value))]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

def CreateCsv(companymail, chanid, videoid, videotitle, videoDuration, numberofviews, publishdate, links):
    companyname = domain_name(companymail)
    channelURL = 'https://youtube.com/channel/' + chanid
    videoURL = 'https://www.youtube.com/watch?v=' + videoid
    wb = load_workbook('SentEmails.xlsx')
    ws = wb.worksheets[0]
    counter = 1;
    for row in ws.iter_rows(min_row = 2):
        counter += 1
    now = datetime.now()
    now = now.strftime("%d%m%Y %H:%M:%S")
    data = [str(counter) + '/' + now, companyname, companymail, channelURL, videoURL, videotitle, videoDuration, numberofviews, publishdate, links]
    ws.append(data)
    expandCsvCells(ws)
    wb.save('SentEmails.xlsx')

# Function to test if the values are corect
def numbertransform(x):
    try:
        a = int(x.get())
        return a
    except:
        labelerror = customtkinter.CTkLabel(master=frame_1, text='Invalid data.')
        labelerror.pack(side = tkinter.BOTTOM)
        return ValueError

# Function to extract the domain name from the URL in order to search it on youtube
def domain_name(url):
    return re.search('(https?://)?(www\d?\.)?(?P<name>[\w-]+)\.', url).group('name')

# Function to send the email to the company
def SendEmail(companywebsite, channeltitle, videotitle, numberofviews, numberoflikes, email):
    companyname = domain_name(companywebsite)
    data = open('index.html', 'r').read()
    temp = Template(data)
    message = Mail(
        from_email='email from SendGrid',
        to_emails= email,
        subject='Promote your page right now!',
        html_content = temp.render(channeltitle = channeltitle, videotitle = videotitle, numberofviews = numberofviews, numberoflikes = numberoflikes)
    )
    try:
        sg = SendGridAPIClient('SAPI KEY')
        response = sg.send(message)
        print(response.status_code)
    except Exception as e:
        print(e.message)

def YoutubeSearch():
    global links
    # this is used to build our youtube api
    youtube = build('youtube', 'v3', developerKey= 'YOUR API KEY')
    date_tolook = CalculateDate() #Function to calculate the date to look for
    for item in ls:
        request = youtube.search().list(
            type = "channel",
            part = "snippet",
            q = domain_name(item),
            maxResults = 1
        )
        response = request.execute()
        if response['pageInfo']['totalResults'] > 0:
            chanid = response['items'][0]['id']['channelId']
            request2 = youtube.search().list(
                part="snippet",
                type = "video",
                channelId= chanid,
                publishedAfter= date_tolook,
                videoDuration=videoentry,
                order = "date"
            )
            response2 = request2.execute()
            if response2['pageInfo']['totalResults'] > 0:
                title = response2['items'][0]['snippet']['title']
                videoid = response2['items'][0]['id']['videoId']
                channeltitle = response2['items'][0]['snippet']['title']
                publishdate = response2['items'][0]['snippet']['publishedAt']
                publishdate = publishdate[:10] + publishdate[11:18]
                request3 = youtube.videos().list(
                    part="contentDetails, statistics",
                    id = videoid
                )
                response3 = request3.execute()
                numberofviews = int(response3['items'][0]['statistics']['viewCount'])
                numberoflikes = response3['items'][0]['statistics']['likeCount']
                videoDuration = response3['items'][0]['contentDetails']['duration']
                videoDurationSubString = re.findall(r'\d+', videoDuration)
                videoDuration = videoDurationSubString[0] + ':' + videoDurationSubString[1]
                # This will scrap the youtube channel for links.
                content = urllib.request.urlopen('https://www.youtube.com/channel/' + chanid + '/about').read().decode('utf-8');
                parts = content.split('u0026q=')
                partsLen = len(parts)
                links = []
                for partsIndex in range(partsLen):
                    part = parts[partsIndex]
                    realPart = part.split('\"')[0]
                    if realPart[:4] == 'http':
                        realPart = realPart.replace('%3A', ':').replace('%2F', '/')
                        if not realPart in links:
                            links += [realPart]
                SendEmail(item, title, channeltitle, numberofviews, numberoflikes, emaills[ls.index(item)])
                if not links:
                    links = "No social media links found"
                else: links = ' '.join(links)
                CreateCsv(item, chanid, videoid, title, videoDuration, numberofviews, publishdate, links)
                done = customtkinter.CTkLabel(master=frame_1, text='Sent {} mails'.format(numberofmails))
                done.pack(side = tkinter.TOP)

# Function for extracting the data when pressing the button 'Submit'. This will transform the data into a readeable form
# it will test if the data entered is integer then will pass to youtube search
def doSomething():
    global videoentry, maxdate, done, labelerror, numberofmails
    numberofmails = 0
    # Testing if the value entered is ok. Video length and date must be int types
    try:
        if done.winfo_exists() == 1:
           done.pack_forget()
    except: pass
    try: 
        videoentry = entry1.get()
        if videoentry == "Long(Longer than 20 mins)":
            videoentry = "long"
        elif videoentry == "Medium(Between 4 and 20 mins)":
            videoentry = "medium"
        elif videoentry == "Short(Less than 4 mins)":
            videoentry = "short"
        else: videoentry = "any"
        maxdate = int(entry2.get())
    except:
        labelerror = customtkinter.CTkLabel(master=frame_1, text='Invalid data.')
        labelerror.pack(side = tkinter.TOP)
        raise ValueError
    try:
        if labelerror.winfo_exists() == 1:
           labelerror.pack_forget()
    except: pass
    YoutubeSearch()

def uploadCSV():
    # This is the list where we will extract all the names of the companies
    global ls, emaills
    ls = []
    emaills = []
    # We will use panda library to read the csv and extract the data
    f_types = [('CSV files',"*.csv"),('All',"*.*")]
    file = tkinter.filedialog.askopenfilename(filetypes=f_types)
    df=pd.read_csv(file) # creating panda DataFrame
    for ind in df.index:
        ls.append(df['Website'][ind])
        emaills.append(df['EmailAddress'][ind])
    ok=customtkinter.CTkLabel(app, text= file + " uploaded")
    ok.pack(pady=6, padx=10)

def CreateInterface():
    global app, entry1, entry2, frame_1
    customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
    customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green
    app = customtkinter.CTk()
    app.geometry("520x400")
    app.title("Marketing Bot")
    app.iconbitmap("terminator.ico")
    frame_1 = customtkinter.CTkFrame(master=app)
    frame_1.pack(pady=20, padx=60, fill="both", expand=True)

    label_0 = customtkinter.CTkLabel(master=frame_1, text='Insert the company names CSV file:', justify=tkinter.LEFT)
    b0 = customtkinter.CTkButton(master=frame_1, text='Browse File', width=10,command = lambda:uploadCSV())
    label_0.pack(side = tkinter.TOP)
    b0.pack(pady=6, padx=10)

    label_1 = customtkinter.CTkLabel(master=frame_1, text="Video duration:", justify=tkinter.LEFT)
    label_1.pack(side = tkinter.TOP)
    vlist = ["any", 'Long(Longer than 20 mins)', "Medium(Between 4 and 20 mins)", "Short(Less than 4 mins)"]
    entryvalue1 = tkinter.StringVar()
    entry1 = customtkinter.CTkComboBox(master = frame_1, values = vlist, variable = entryvalue1)
    entry1['state'] = 'readonly'
    entry1.set("Pick an Option")
    entry1.pack(pady = 6, padx = 10)

    label_2 = customtkinter.CTkLabel(master=frame_1, text="Posted in the last(days):", justify=tkinter.LEFT)
    entryvalue2 = tkinter.StringVar()
    entry2 = customtkinter.CTkEntry(master=frame_1, textvariable=entryvalue2, placeholder_text="Enter a number value")
    label_2.pack(side = tkinter.TOP)
    entry2.pack(pady=6, padx=10)

    button1 = customtkinter.CTkButton(master=frame_1, text="Submit", command=doSomething)
    button2 = customtkinter.CTkButton(master=frame_1, text='Quit', command=app.quit)
    button1.pack(pady=6, padx=10)
    button2.pack(pady=6, padx=10)

    app.mainloop()

if __name__ == "__main__":
    CreateInterface()
